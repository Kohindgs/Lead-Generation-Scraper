"""
CRM Export & Reporting Module
==============================
Exports leads and outreach to:
  1. Excel / CSV  — ready for import to any CRM (HubSpot, Salesforce, Pipedrive)
  2. JSON — for API integrations
  3. HTML report — visual dashboard with stats, charts, top leads

Also generates a daily / per-campaign PDF summary report.
"""
import json
import csv
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from src.models import Lead, OutreachMessage, ScrapingResult
from src.utils.database import get_leads, get_stats
from src.utils.helpers import get_logger

logger = get_logger(__name__)

EXPORT_DIR = Path(__file__).resolve().parent.parent.parent / "exports"
REPORT_DIR = Path(__file__).resolve().parent.parent.parent / "reports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)
REPORT_DIR.mkdir(parents=True, exist_ok=True)


# ── Excel / CSV export ────────────────────────────────────────────────────────

LEAD_FIELDS = [
    "id", "full_name", "first_name", "last_name", "title", "seniority",
    "company_name", "company_website", "company_size", "industry",
    "email", "email_verified", "phone", "linkedin_url",
    "city", "state", "country",
    "google_rating", "google_review_count",
    "lead_score", "icp_match", "status", "source",
    "pain_points", "services_needed", "buying_signals",
    "scraped_at", "notes", "tags",
]


def export_to_csv(leads: List[Lead], filename: Optional[str] = None) -> Path:
    """Export leads to a CSV file."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = filename or f"leads_{ts}.csv"
    path = EXPORT_DIR / filename

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=LEAD_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for lead in leads:
            row = lead.model_dump()
            # Flatten lists to semicolon-separated strings
            for list_field in ["pain_points", "services_needed", "buying_signals", "tags"]:
                row[list_field] = "; ".join(row.get(list_field) or [])
            row["icp_match"] = "Yes" if row.get("icp_match") else "No"
            row["email_verified"] = "Yes" if row.get("email_verified") else "No"
            row["source"] = row.get("source") or ""
            row["status"] = row.get("status") or ""
            if row.get("scraped_at"):
                row["scraped_at"] = str(row["scraped_at"])
            writer.writerow({k: row.get(k, "") for k in LEAD_FIELDS})

    logger.info("Exported %d leads to CSV: %s", len(leads), path)
    return path


def export_to_excel(leads: List[Lead], filename: Optional[str] = None) -> Path:
    """Export leads to Excel with formatting."""
    try:
        import pandas as pd
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("pandas/openpyxl not installed — falling back to CSV.")
        return export_to_csv(leads, (filename or "leads").replace(".xlsx", ".csv"))

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = filename or f"leads_{ts}.xlsx"
    path = EXPORT_DIR / filename

    rows = []
    for lead in leads:
        d = lead.model_dump()
        rows.append({
            "Score": d.get("lead_score", 0),
            "ICP Match": "✓" if d.get("icp_match") else "",
            "Full Name": d.get("full_name", ""),
            "First Name": d.get("first_name", ""),
            "Last Name": d.get("last_name", ""),
            "Title": d.get("title", ""),
            "Company": d.get("company_name", ""),
            "Industry": d.get("industry", ""),
            "Website": d.get("company_website", ""),
            "Email": d.get("email", ""),
            "Email Verified": "Yes" if d.get("email_verified") else "No",
            "Phone": d.get("phone", ""),
            "LinkedIn": d.get("linkedin_url", ""),
            "City": d.get("city", ""),
            "Country": d.get("country", ""),
            "Google Rating": d.get("google_rating", ""),
            "Reviews": d.get("google_review_count", ""),
            "Pain Points": "; ".join(d.get("pain_points") or []),
            "Services Needed": "; ".join(d.get("services_needed") or []),
            "Buying Signals": "; ".join(d.get("buying_signals") or []),
            "Status": d.get("status", "new"),
            "Source": d.get("source", ""),
            "Scraped At": str(d.get("scraped_at", "")),
        })

    df = pd.DataFrame(rows)
    df = df.sort_values("Score", ascending=False)

    with pd.ExcelWriter(str(path), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Leads", index=False)
        wb = writer.book
        ws = writer.sheets["Leads"]

        # Header styling
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Score column colouring (column A = Score)
        for row_idx in range(2, len(rows) + 2):
            score_cell = ws.cell(row=row_idx, column=1)
            score = score_cell.value or 0
            if score >= 75:
                score_cell.fill = PatternFill("solid", fgColor="C6EFCE")  # green
            elif score >= 55:
                score_cell.fill = PatternFill("solid", fgColor="FFEB9C")  # yellow
            else:
                score_cell.fill = PatternFill("solid", fgColor="FFC7CE")  # red

        # Auto-width columns
        for col_idx, col in enumerate(df.columns, 1):
            max_len = max(
                df[col].astype(str).map(len).max(),
                len(col)
            ) + 2
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 40)

    logger.info("Exported %d leads to Excel: %s", len(leads), path)
    return path


def export_to_json(leads: List[Lead], filename: Optional[str] = None) -> Path:
    """Export leads to JSON."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = filename or f"leads_{ts}.json"
    path = EXPORT_DIR / filename

    data = [json.loads(lead.model_dump_json()) for lead in leads]
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)

    logger.info("Exported %d leads to JSON: %s", len(leads), path)
    return path


def export_outreach_messages(
    leads: List[Lead],
    messages: List[OutreachMessage],
    filename: Optional[str] = None,
) -> Path:
    """Export outreach messages as a CSV for mail-merge / CRM import."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = filename or f"outreach_{ts}.csv"
    path = EXPORT_DIR / filename

    lead_map = {l.id: l for l in leads if l.id}

    with open(path, "w", newline="", encoding="utf-8") as f:
        fields = [
            "lead_id", "full_name", "company_name", "email",
            "linkedin_url", "channel", "subject",
            "message", "follow_up_1", "follow_up_2", "follow_up_3",
        ]
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for msg in messages:
            lead = lead_map.get(msg.lead_id)
            writer.writerow({
                "lead_id": msg.lead_id,
                "full_name": lead.full_name if lead else "",
                "company_name": lead.company_name if lead else "",
                "email": lead.email if lead else "",
                "linkedin_url": lead.linkedin_url if lead else "",
                "channel": msg.channel.value,
                "subject": msg.subject or "",
                "message": msg.message,
                "follow_up_1": msg.follow_up_1 or "",
                "follow_up_2": msg.follow_up_2 or "",
                "follow_up_3": msg.follow_up_3 or "",
            })

    logger.info("Exported %d outreach messages: %s", len(messages), path)
    return path


# ── HTML Dashboard Report ─────────────────────────────────────────────────────

def generate_html_report(
    result: ScrapingResult,
    leads: List[Lead],
    messages: List[OutreachMessage],
) -> Path:
    """Generate a visual HTML dashboard report for the campaign run."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    campaign_slug = result.campaign_name.lower().replace(" ", "_")
    path = REPORT_DIR / f"report_{campaign_slug}_{ts}.html"

    stats = get_stats()
    top_leads = sorted(leads, key=lambda l: l.lead_score, reverse=True)[:10]
    icp_count = sum(1 for l in leads if l.icp_match)

    # Industry breakdown
    industry_counts: dict = {}
    for lead in leads:
        ind = lead.industry or "Unknown"
        industry_counts[ind] = industry_counts.get(ind, 0) + 1
    top_industries = sorted(industry_counts.items(), key=lambda x: x[1], reverse=True)[:8]

    # Source breakdown
    source_counts: dict = {}
    for lead in leads:
        src = lead.source.value
        source_counts[src] = source_counts.get(src, 0) + 1

    html = _render_html_report(
        result, leads, top_leads, messages,
        icp_count, top_industries, source_counts, stats
    )

    path.write_text(html, encoding="utf-8")
    logger.info("HTML report generated: %s", path)
    return path


def _render_html_report(
    result: ScrapingResult,
    leads: List[Lead],
    top_leads: List[Lead],
    messages: List[OutreachMessage],
    icp_count: int,
    top_industries: list,
    source_counts: dict,
    stats: dict,
) -> str:
    from src.config import agency

    avg_score = (
        sum(l.lead_score for l in leads) / len(leads) if leads else 0
    )

    top_leads_rows = ""
    for lead in top_leads:
        score_color = (
            "#28a745" if lead.lead_score >= 75
            else "#ffc107" if lead.lead_score >= 55
            else "#dc3545"
        )
        pain = "; ".join(lead.pain_points[:2]) or "—"
        services = ", ".join(lead.services_needed[:2]) or "—"
        linkedin_link = (
            f'<a href="{lead.linkedin_url}" target="_blank">LinkedIn</a>'
            if lead.linkedin_url else "—"
        )
        top_leads_rows += f"""
        <tr>
            <td><b>{lead.full_name or lead.company_name}</b></td>
            <td>{lead.title or "—"}</td>
            <td>{lead.company_name or "—"}</td>
            <td>{lead.industry or "—"}</td>
            <td>{lead.email or "—"}</td>
            <td>{linkedin_link}</td>
            <td style="color:{score_color};font-weight:bold;">{lead.lead_score}</td>
            <td style="font-size:0.85em;">{pain}</td>
            <td style="font-size:0.85em;">{services}</td>
        </tr>"""

    industry_bars = ""
    max_count = top_industries[0][1] if top_industries else 1
    for ind, cnt in top_industries:
        pct = int(cnt / max_count * 100)
        industry_bars += f"""
        <div style="margin-bottom:8px;">
            <div style="display:flex;align-items:center;gap:10px;">
                <span style="width:180px;font-size:0.9em;">{ind}</span>
                <div style="flex:1;background:#eee;border-radius:4px;height:18px;">
                    <div style="width:{pct}%;background:#1F4E79;height:18px;border-radius:4px;"></div>
                </div>
                <span style="width:30px;text-align:right;font-size:0.9em;">{cnt}</span>
            </div>
        </div>"""

    source_badges = ""
    source_colors = {
        "linkedin": "#0077B5",
        "google_maps": "#34A853",
        "google_search": "#EA4335",
        "manual": "#6c757d",
    }
    for src, cnt in source_counts.items():
        color = source_colors.get(src, "#6c757d")
        source_badges += (
            f'<span style="background:{color};color:#fff;padding:6px 14px;'
            f'border-radius:20px;margin:4px;display:inline-block;">'
            f'{src.replace("_", " ").title()}: <b>{cnt}</b></span>'
        )

    duration = (
        f"{result.duration_seconds:.0f}s" if result.duration_seconds else "—"
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{agency.name} – Lead Generation Report</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', Arial, sans-serif; background: #f4f6f9; color: #333; }}
  .header {{ background: linear-gradient(135deg, #1F4E79, #2E86AB); color: #fff;
             padding: 32px 40px; }}
  .header h1 {{ font-size: 2em; margin-bottom: 6px; }}
  .header p {{ opacity: 0.85; font-size: 1.05em; }}
  .container {{ max-width: 1200px; margin: 30px auto; padding: 0 20px; }}
  .stats-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
                 gap: 16px; margin-bottom: 28px; }}
  .stat-card {{ background: #fff; border-radius: 10px; padding: 20px;
               box-shadow: 0 2px 8px rgba(0,0,0,0.07); text-align: center; }}
  .stat-card .value {{ font-size: 2.2em; font-weight: bold; color: #1F4E79; }}
  .stat-card .label {{ font-size: 0.9em; color: #888; margin-top: 4px; }}
  .section {{ background: #fff; border-radius: 10px; padding: 24px;
              box-shadow: 0 2px 8px rgba(0,0,0,0.07); margin-bottom: 24px; }}
  .section h2 {{ font-size: 1.2em; color: #1F4E79; margin-bottom: 16px;
                 border-bottom: 2px solid #e8edf3; padding-bottom: 8px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 0.88em; }}
  th {{ background: #1F4E79; color: #fff; padding: 10px 12px; text-align: left; }}
  td {{ padding: 9px 12px; border-bottom: 1px solid #eee; vertical-align: top; }}
  tr:hover td {{ background: #f8fafc; }}
  .badge {{ background: #e8f4fd; color: #1F4E79; padding: 3px 8px;
            border-radius: 12px; font-size: 0.8em; }}
  footer {{ text-align: center; color: #aaa; padding: 20px; font-size: 0.85em; }}
</style>
</head>
<body>
<div class="header">
  <h1>📊 Lead Generation Report</h1>
  <p>{agency.name} — {agency.website} &nbsp;|&nbsp; Campaign: <b>{result.campaign_name}</b>
     &nbsp;|&nbsp; Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}</p>
</div>

<div class="container">
  <!-- KPI Cards -->
  <div class="stats-grid">
    <div class="stat-card">
      <div class="value">{result.total_scraped}</div>
      <div class="label">Leads Scraped</div>
    </div>
    <div class="stat-card">
      <div class="value">{icp_count}</div>
      <div class="label">ICP Matches</div>
    </div>
    <div class="stat-card">
      <div class="value">{avg_score:.0f}</div>
      <div class="label">Avg Lead Score</div>
    </div>
    <div class="stat-card">
      <div class="value">{result.total_outreach_generated}</div>
      <div class="label">Outreach Generated</div>
    </div>
    <div class="stat-card">
      <div class="value">{stats.get('total_leads', 0)}</div>
      <div class="label">Total DB Leads</div>
    </div>
    <div class="stat-card">
      <div class="value">{duration}</div>
      <div class="label">Run Duration</div>
    </div>
  </div>

  <!-- Sources -->
  <div class="section">
    <h2>Lead Sources</h2>
    {source_badges}
  </div>

  <!-- Industry Breakdown -->
  <div class="section">
    <h2>Top Industries</h2>
    {industry_bars}
  </div>

  <!-- Top Leads Table -->
  <div class="section">
    <h2>Top 10 Leads by Score</h2>
    <table>
      <thead>
        <tr>
          <th>Name</th><th>Title</th><th>Company</th><th>Industry</th>
          <th>Email</th><th>LinkedIn</th><th>Score</th>
          <th>Pain Points</th><th>Services Needed</th>
        </tr>
      </thead>
      <tbody>{top_leads_rows}</tbody>
    </table>
  </div>

  <!-- Errors -->
  {f'<div class="section"><h2>Errors ({len(result.errors)})</h2><ul>' +
    "".join(f"<li>{e}</li>" for e in result.errors) + "</ul></div>"
    if result.errors else ""}
</div>

<footer>
  Generated by DGenius Lead Generation System &bull; {agency.name} &bull; {agency.website}
</footer>
</body>
</html>"""
